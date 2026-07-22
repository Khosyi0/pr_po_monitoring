import streamlit as st
import pandas as pd
import plotly.express as px
import io
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.text import Font as DrawingFont, RegularTextRun
from openpyxl.chart.text import Text
from openpyxl.chart.title import Title
from utils import MAPPING_SINGKATAN


def generate_excel_export(df_plot, df_pivot, kolom_tanggal, y_col, y_label, jenis_harga, warna_map, list_resume):
    wb = Workbook()
    ws = wb.active
    ws.title = "Komparasi Harga NH4Cl"
    ws_data = wb.create_sheet("_DataChart")

    df_chart = df_plot.pivot_table(
        index='tanggal_terbit', columns='label_komparasi', values=y_col, aggfunc='mean'
    ).sort_index()

    df_chart_reset = df_chart.reset_index()
    df_chart_reset['tanggal_terbit'] = df_chart_reset['tanggal_terbit'].dt.strftime('%d %b %Y')

    headers = ['Tanggal Terbit'] + list(df_chart.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(df_chart_reset.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    n_rows = len(df_chart_reset)
    n_cols = len(headers)

    # =========================================================================
    # SETUP FONT ARIAL UNTUK CHART
    # =========================================================================
    arial_font = DrawingFont(typeface='Arial')
    cp_arial = CharacterProperties(latin=arial_font)
    cp_arial_bold = CharacterProperties(latin=arial_font, b=True)
    cp_arial_sz700 = CharacterProperties(sz=700, latin=arial_font)

    # Fungsi bantu untuk membungkus teks menjadi Title Chart yang terformat
    def create_formatted_title(text_val, is_bold=True):
        cp = cp_arial_bold if is_bold else cp_arial
        run = RegularTextRun(t=text_val, rPr=cp)
        p = Paragraph(pPr=ParagraphProperties(defRPr=cp), r=[run])
        return Title(tx=Text(rich=RichText(p=[p])))
    # =========================================================================

    chart = LineChart()
    chart.title = create_formatted_title(f"Komparasi Tren Harga NH4Cl ({jenis_harga})", is_bold=True)
    chart.height = 14
    chart.width = 32
    chart.style = None
    chart.title.overlay = False

    data_ref = Reference(ws_data, min_col=2, max_col=n_cols, min_row=1, max_row=n_rows + 1)
    cats_ref = Reference(ws_data, min_col=1, max_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(data_ref, titles_from_data=True, from_rows=False)
    chart.set_categories(cats_ref)

    label_columns = list(df_chart.columns)
    for series, label in zip(chart.series, label_columns):
        series.marker.symbol = "none"
        series.smooth = False
        hex_color = warna_map.get(label, "#1f77b4").lstrip('#').upper()
        series.graphicalProperties.line.width = 18000
        series.graphicalProperties.line.solidFill = hex_color

    # 2. Terapkan Arial pada Axis Y (Judul dan Label Angka)
    chart.y_axis.title = create_formatted_title(y_label, is_bold=True)
    chart.y_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_arial), endParaRPr=cp_arial)]
    )
    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.graphicalProperties = GraphicalProperties()
    chart.y_axis.majorGridlines.graphicalProperties.line = LineProperties(solidFill="E0E0E0", w=9525)
    chart.y_axis.delete = False

    # 3. Terapkan Arial pada Axis X (Judul dan Label Tanggal rotasi)
    chart.x_axis.title = create_formatted_title("Tanggal Publikasi", is_bold=True)
    chart.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=-5400000, vert="horz"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_arial_sz700),
                     endParaRPr=cp_arial_sz700)]
    )
    chart.x_axis.delete = False
    chart.x_axis.majorGridlines = None
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1

    # 4. Terapkan Arial pada Legend
    chart.legend.position = 'b'
    chart.legend.overlay = False
    chart.legend.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_arial), endParaRPr=cp_arial)]
    )

    chart.layout = Layout(
        manualLayout=ManualLayout(x=0.02, y=0.18, h=0.64, w=0.90, xMode="edge", yMode="edge")
    )

    ws.add_chart(chart, "A1")

    # ========================== STYLING TABEL ==========================
    HEADER_BLUE = "BDD7EE"
    thin = Side(style='thin', color='000000')
    TABLE_START_ROW = 34
    n_date_cols = len(kolom_tanggal)
    last_col = 1 + n_date_cols

    ws.cell(row=TABLE_START_ROW, column=1,
            value="Detail Histori Data (3 Periode Terakhir)").font = Font(bold=True, size=13)

    header_row1 = TABLE_START_ROW + 1
    header_row2 = header_row1 + 1
    first_data_row = header_row2 + 1
    last_data_row = first_data_row + len(df_pivot) - 1

    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=header_row1, column=col_idx)
        cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=header_row1, column=1, value="Referensi")
    ws.cell(row=header_row1, column=2, value="Harga USD/MT")

    ws.cell(row=header_row2, column=1).fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    for col_idx, tgl_label in enumerate(kolom_tanggal, start=2):
        cell = ws.cell(row=header_row2, column=col_idx, value=tgl_label)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, (index_label, row) in enumerate(df_pivot.iterrows()):
        r = first_data_row + row_offset
        cell = ws.cell(row=r, column=1, value=index_label)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx, col in enumerate(df_pivot.columns, start=2):
            val = row[col]
            val_str = "" if pd.isna(val) else str(val)
            c = ws.cell(row=r, column=col_idx, value=val_str)
            c.alignment = Alignment(horizontal="center", vertical="center")

    border_thin_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(header_row1, last_data_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).border = border_thin_all

    ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)
    ws.merge_cells(start_row=header_row1, start_column=2, end_row=header_row1, end_column=last_col)

    resume_title_row = last_data_row + 2
    ws.cell(row=resume_title_row, column=1, value="Resume :").font = Font(bold=True, size=11, italic=True, name='Arial')
    
    for idx, poin in enumerate(list_resume, start=1):
        current_resume_row = resume_title_row + idx
        cell = ws.cell(row=current_resume_row, column=1, value=f"•  {poin}")
        cell.font = Font(size=11, name='Arial')
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        ws.merge_cells(start_row=current_resume_row, start_column=1, end_row=current_resume_row, end_column=last_col)
        
        # Kalkulasi tinggi baris otomatis
        jumlah_baris = (len(poin) // 90) + 1
        ws.row_dimensions[current_resume_row].height = 16 * jumlah_baris

    ws.column_dimensions['A'].width = 26
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    ws.row_dimensions[header_row1].height = 22
    ws.row_dimensions[header_row2].height = 20

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    is_bold = cell.font.bold if cell.font else False
                    cell.font = Font(name='Arial', size=11, bold=is_bold)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def variasikan_warna(hex_color, index, total):
    if total <= 1:
        return hex_color
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    factor = 0.6 + (0.7 * index / max(total - 1, 1))
    r = min(255, int(r * factor))
    g = min(255, int(g * factor))
    b = min(255, int(b * factor))
    return f"#{r:02x}{g:02x}{b:02x}"


def hitung_resume_nh4cl(df_plot, y_col):
    bulan_indo = {1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
                  7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'}

    def _get_nama_minggu(dt):
        if dt.day <= 7: return f"awal {bulan_indo[dt.month]} {dt.year}"
        elif dt.day <= 14: return f"minggu kedua {bulan_indo[dt.month]} {dt.year}"
        elif dt.day <= 21: return f"minggu ketiga {bulan_indo[dt.month]} {dt.year}"
        else: return f"akhir {bulan_indo[dt.month]} {dt.year}"

    if df_plot.empty: return ["Data tidak tersedia."]

    tgl_T0 = pd.Timestamp(df_plot['tanggal_terbit'].max())
    batas_1_bulan = tgl_T0 - pd.DateOffset(months=1)
    batas_2_bulan = tgl_T0 - pd.DateOffset(months=2)

    df_T0 = df_plot.sort_values('tanggal_terbit').drop_duplicates(subset=['label_komparasi'], keep='last')
    harga_T0 = df_T0[y_col].mean()

    df_T1_range = df_plot[(df_plot['tanggal_terbit'] >= batas_1_bulan) & (df_plot['tanggal_terbit'] < tgl_T0)]
    harga_T1 = df_T1_range.sort_values('tanggal_terbit').drop_duplicates(subset=['label_komparasi'], keep='last')[y_col].mean() if not df_T1_range.empty else harga_T0
    tgl_T1 = pd.Timestamp(df_T1_range['tanggal_terbit'].max()) if not df_T1_range.empty else tgl_T0

    df_T2_range = df_plot[(df_plot['tanggal_terbit'] >= batas_2_bulan) & (df_plot['tanggal_terbit'] < batas_1_bulan)]
    harga_T2 = df_T2_range.sort_values('tanggal_terbit').drop_duplicates(subset=['label_komparasi'], keep='first')[y_col].mean() if not df_T2_range.empty else harga_T0
    tgl_T2 = pd.Timestamp(df_T2_range['tanggal_terbit'].min()) if not df_T2_range.empty else tgl_T0

    delta_recent = harga_T0 - harga_T1
    delta_past = harga_T1 - harga_T2
    threshold = 10.0 

    tren = "menurun" if delta_recent < -2.0 else ("meningkat" if delta_recent > 2.0 else "stabil")
    signifikansi = " signifikan" if abs(delta_recent) >= threshold else " tidak signifikan"
    
    poin_1 = f"Secara keseluruhan, harga NH4Cl menunjukkan tren {tren} ({signifikansi}) pada {_get_nama_minggu(tgl_T0.date())}."
    poin_2 = f"Harga saat ini (USD {harga_T0:.2f}/MT) terpaut {abs(harga_T0 - harga_T2):.2f} USD/MT jika dibandingkan dengan periode {_get_nama_minggu(tgl_T2.date())}."

    list_resume = [poin_1, poin_2]
    # Deteksi mandek
    for label in df_plot['label_komparasi'].unique():
        if (tgl_T0 - pd.Timestamp(df_plot[df_plot['label_komparasi'] == label]['tanggal_terbit'].max())).days > 14:
            list_resume.append(f"Untuk referensi {label}, harga terakhir dirilis pada tanggal lama.")
    return list_resume


def render(load_data, global_context):
    st.markdown("### :material/science: Analisis Tren Komparasi Harga Pasar: NH4Cl")

    from config_db import get_setting
    from datetime import datetime

    bahan_baku_date_str = get_setting("DATA_UPDATE_BAHAN_BAKU", "2026-03-31")
    try:
        tgl_update_bb = datetime.strptime(bahan_baku_date_str, "%Y-%m-%d").date()
    except:
        tgl_update_bb = datetime(2026, 3, 31).date()

    bulan_indo_header = {
        1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
        7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
    }
    tgl_update_str = f"{tgl_update_bb.day:02d} {bulan_indo_header[tgl_update_bb.month]} {tgl_update_bb.year}"

    st.markdown(
        f"<p style='font-size:14px; opacity:0.65; margin-top:-6px; margin-bottom:16px;'>"
        f"Data terakhir diperbarui pada <b>{tgl_update_str}</b>"
        f"</p>",
        unsafe_allow_html=True
    )

    query = """
        SELECT tanggal_terbit, nama_majalah, incoterm, harga_min, harga_max 
        FROM master_harga_bahan_baku 
        WHERE bahan_baku = 'NH4Cl'
        ORDER BY tanggal_terbit ASC
    """
    df = load_data(query)

    if df.empty:
        st.warning("Data harga NH4Cl belum tersedia di database.")
        return

    list_majalah = df['nama_majalah'].unique()
    min_date = df['tanggal_terbit'].min()
    max_date = df['tanggal_terbit'].max()

    default_start_date = pd.Timestamp('2025-01-01').date()
    calendar_min_date = min(min_date, default_start_date)

    if default_start_date > max_date or default_start_date < min_date:
        default_start_date = min_date

    def _save_to_permanent(widget_key, permanent_key):
        st.session_state[permanent_key] = st.session_state[widget_key]

    with st.expander(":material/settings: Filter Komparasi Harga Pasar", expanded=True):
        col_mulai, col_sampai, col_metode, col_jml = st.columns(4)
        with col_mulai:
            start_date = st.date_input(
                "Mulai dari tanggal",
                value=st.session_state.get("_perm_start_date_nh4cl", default_start_date),
                min_value=calendar_min_date,
                max_value=max_date,
                key="start_date_nh4cl",
                on_change=_save_to_permanent,
                args=("start_date_nh4cl", "_perm_start_date_nh4cl")
            )
        with col_sampai:
            end_date = st.date_input(
                "Sampai tanggal",
                value=st.session_state.get("_perm_end_date_nh4cl", max_date),
                min_value=calendar_min_date,
                max_value=max_date,
                key="end_date_nh4cl",
                on_change=_save_to_permanent,
                args=("end_date_nh4cl", "_perm_end_date_nh4cl")
            )
        with col_metode:
            jenis_harga_options = ["AVERAGE", "MIN", "MAX"]
            jenis_harga_default = st.session_state.get("_perm_jenis_harga_nh4cl", "AVERAGE")
            jenis_harga = st.selectbox(
                "Jenis Harga",
                jenis_harga_options,
                index=jenis_harga_options.index(jenis_harga_default) if jenis_harga_default in jenis_harga_options else 0,
                help="Pilih nilai harga yang ingin diplot pada grafik",
                key="jenis_harga_nh4cl",
                on_change=_save_to_permanent,
                args=("jenis_harga_nh4cl", "_perm_jenis_harga_nh4cl")
            )
        with col_jml:
            jml_komparasi = st.number_input(
                "Jumlah Komparasi",
                min_value=1, max_value=5,
                value=st.session_state.get("_perm_jml_komparasi_nh4cl", 2),
                key="jml_komparasi_nh4cl",
                on_change=_save_to_permanent,
                args=("jml_komparasi_nh4cl", "_perm_jml_komparasi_nh4cl")
            )

        st.markdown("<hr style='margin: 10px 0; border-color: rgba(255,255,255,0.1);'>", unsafe_allow_html=True)

        komparasi_data = []
        warna_map = {}
        default_colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd"]

        for i in range(int(jml_komparasi)):
            c1, c2, c3 = st.columns([3, 3, 1])
            with c1:
                perm_key_majalah = f"_perm_majalah_nh4cl_{i}"
                default_majalah = st.session_state.get(perm_key_majalah, list_majalah[i] if i < len(list_majalah) else list_majalah[0])
                majalah_index = list(list_majalah).index(default_majalah) if default_majalah in list_majalah else 0
                majalah_pilihan = st.selectbox(
                    f"Majalah ke-{i+1}", list_majalah,
                    index=majalah_index,
                    key=f"majalah_nh4cl_{i}",
                    on_change=_save_to_permanent,
                    args=(f"majalah_nh4cl_{i}", perm_key_majalah)
                )
            with c2:
                list_incoterm = df[df['nama_majalah'] == majalah_pilihan]['incoterm'].unique()
                perm_key_incoterm = f"_perm_incoterm_nh4cl_{i}"
                default_incoterm = st.session_state.get(perm_key_incoterm, list_incoterm[0] if len(list_incoterm) > 0 else None)
                incoterm_index = list(list_incoterm).index(default_incoterm) if default_incoterm in list_incoterm else 0
                incoterm_pilihan = st.selectbox(
                    f"Metode Incoterm ke-{i+1}", list_incoterm,
                    index=incoterm_index if len(list_incoterm) > 0 else None,
                    key=f"incoterm_nh4cl_{i}",
                    on_change=_save_to_permanent,
                    args=(f"incoterm_nh4cl_{i}", perm_key_incoterm)
                )
            with c3:
                perm_key_warna = f"_perm_color_nh4cl_{i}"
                default_warna = st.session_state.get(perm_key_warna, default_colors[i % len(default_colors)])
                warna_pilihan = st.color_picker(
                    "Warna", default_warna,
                    key=f"color_nh4cl_{i}",
                    on_change=_save_to_permanent,
                    args=(f"color_nh4cl_{i}", perm_key_warna)
                )

            if incoterm_pilihan:
                komparasi_data.append({
                    "majalah": majalah_pilihan,
                    "incoterms": [incoterm_pilihan],
                    "warna_dasar": warna_pilihan
                })

        for item in komparasi_data:
            for idx, incoterm in enumerate(item["incoterms"]):
                # 1. Gabungkan nama aslinya
                label_asli = f"{item['majalah']} - {incoterm}"
                
                # 2. Ubah menjadi singkatan agar sinkron dengan yang ada di df_plot nanti
                label_singkat = MAPPING_SINGKATAN.get(label_asli, label_asli)
                
                # 3. Masukkan ke warna_map menggunakan label yang sudah disingkat
                warna_final = variasikan_warna(item["warna_dasar"], idx, len(item["incoterms"]))
                warna_map[label_singkat] = warna_final
                label = f"{item['majalah']} - {incoterm}"
                warna_final = variasikan_warna(item["warna_dasar"], idx, len(item["incoterms"]))
                warna_map[label] = warna_final

    if start_date <= end_date and komparasi_data:
        df_plot = pd.DataFrame()

        for item in komparasi_data:
            majalah = item["majalah"]
            incoterms = item["incoterms"]
            temp_df = df[(df['nama_majalah'] == majalah) & (df['incoterm'].isin(incoterms)) &
                         (df['tanggal_terbit'] >= start_date) & (df['tanggal_terbit'] <= end_date)].copy()
            if not temp_df.empty:
                # 1. Buat kolom nama aslinya
                temp_df['label_komparasi'] = temp_df['nama_majalah'] + ' - ' + temp_df['incoterm']
                
                # 2. Timpa namanya menggunakan dictionary dari utils.py
                temp_df['label_komparasi'] = temp_df['label_komparasi'].apply(
                    lambda x: MAPPING_SINGKATAN.get(x, x)
                )
                
                df_plot = pd.concat([df_plot, temp_df], ignore_index=True)
                

        if not df_plot.empty:
            df_plot['harga_avg'] = (df_plot['harga_min'] + df_plot['harga_max']) / 2
            df_plot['tanggal_terbit'] = pd.to_datetime(df_plot['tanggal_terbit'])
            df_plot = df_plot.sort_values('tanggal_terbit')
            tanggal_unik = df_plot['tanggal_terbit'].unique()

            if jenis_harga == "MIN": y_col, y_label = 'harga_min', 'Harga Minimum (USD/MT)'
            elif jenis_harga == "MAX": y_col, y_label = 'harga_max', 'Harga Maksimum (USD/MT)'
            else: y_col, y_label = 'harga_avg', 'Harga Rata-rata (USD/MT)'

            fig = px.line(
                df_plot, x='tanggal_terbit', y=y_col, color='label_komparasi',
                color_discrete_map=warna_map,
                title=f"Komparasi Tren Harga NH4Cl ({jenis_harga})",
                labels={y_col: y_label, 'tanggal_terbit': 'Tanggal Publikasi', 'label_komparasi': 'Majalah & Incoterm'}
            )

            fig.update_layout(
                hovermode="x unified",
                legend=dict(orientation="v", yanchor="top", y=-0.6, xanchor="left", x=0),
                margin=dict(b=300, t=80, l=60, r=40),
                height=600
            )

            fig.update_xaxes(
                tickangle=-90, type='date', tickmode='array', tickvals=tanggal_unik,
                tickformat="%d %b %Y", title=dict(text="Tanggal Publikasi", standoff=40)
            )

            fig.update_yaxes(dtick=50)

            st.plotly_chart(fig, use_container_width=True)

            st.markdown("#### :material/table_chart: Detail Histori Data (3 Periode Terakhir)")

            df_display = df_plot.copy()
            df_display['harga_range'] = df_display['harga_min'].apply(lambda x: f"{x:.2f}") + ' - ' + df_display['harga_max'].apply(lambda x: f"{x:.2f}")

            df_pivot = df_display.pivot_table(
                index='label_komparasi', columns='tanggal_terbit', values='harga_range',
                aggfunc=lambda x: ' '.join(x)
            )

            df_pivot = df_pivot.sort_index(axis=1, ascending=False)
            df_pivot = df_pivot.iloc[:, :3]

            bulan_indo = {
                1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
                7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
            }
            kolom_tanggal = [f"{d.day:02d} {bulan_indo[d.month]} {d.year}" for d in df_pivot.columns]

            jml_kolom = len(kolom_tanggal)

            thead = f'''
<thead>
    <tr>
        <th rowspan="2" style="vertical-align: middle; text-align: left !important;">Referensi</th>
        <th colspan="{jml_kolom}">Harga USD/MT</th>
    </tr>
    <tr>
'''
            for tgl in kolom_tanggal:
                thead += f"<th>{tgl}</th>"
            thead += "</tr>\n</thead>"

            tbody = "<tbody>\n"
            for index, row in df_pivot.iterrows():
                tbody += f"<tr>\n<td style='text-align: left !important;'>{index}</td>\n"
                for col in df_pivot.columns:
                    val = row[col]
                    val_str = "" if pd.isna(val) else str(val)
                    tbody += f"<td>{val_str}</td>\n"
                tbody += "</tr>\n"
            tbody += "</tbody>"

            html_table = f"<table>\n{thead}\n{tbody}\n</table>"

            styled_html = f"""
<style>
.custom-table-container {{
    width: 100%;
    overflow-x: auto;
    margin-bottom: 2rem;
}}
.custom-table-container table {{
    width: 100%;
    border-collapse: collapse;
    font-family: "Source Sans Pro", sans-serif;
    font-size: 14px;
    color: var(--text-color);
}}
.custom-table-container th, .custom-table-container td {{
    text-align: center !important;
    padding: 10px !important;
    border: 1px solid rgba(128, 128, 128, 0.2);
}}
.custom-table-container th {{
    background-color: rgba(128, 128, 128, 0.1);
    font-weight: 600;
}}
</style>
<div class="custom-table-container">
    {html_table}
</div>
"""
            st.markdown(styled_html, unsafe_allow_html=True)

            list_resume_otomatis = hitung_resume_nh4cl(df_plot, y_col)
            st.markdown("##### *Resume :*")
            for poin in list_resume_otomatis:
                st.markdown(f"- {poin}")
            st.markdown("<br>", unsafe_allow_html=True)
            
            excel_buffer = generate_excel_export(
                df_plot=df_plot, df_pivot=df_pivot, kolom_tanggal=kolom_tanggal,
                y_col=y_col, y_label=y_label, jenis_harga=jenis_harga, warna_map=warna_map,
                list_resume=list_resume_otomatis
            )

            st.markdown("""
                <style>
                div[data-testid="stDownloadButton"] button {
                    background-color: #FF4B4B;
                    color: white;
                    border: none;
                }
                div[data-testid="stDownloadButton"] button:hover {
                    background-color: #E54444;
                    color: white;
                    border: none;
                }
                div[data-testid="stDownloadButton"] button:active {
                    background-color: #CE3D3D;
                    color: white;
                }
                </style>
            """, unsafe_allow_html=True)

            st.download_button(
                label=":material/download: Download Excel (Chart + Tabel)",
                data=excel_buffer,
                file_name=f"komparasi_harga_NH4Cl_{jenis_harga}_{start_date}_{end_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            st.info("Tidak ada data yang tersedia untuk kombinasi filter yang dipilih pada rentang waktu tersebut.")
    else:
        if start_date > end_date:
            st.error("❌ 'Mulai dari tanggal' tidak boleh lebih besar dari 'Sampai tanggal'.")
        else:
            st.info("Silakan tentukan minimal 1 metode Incoterm.")
