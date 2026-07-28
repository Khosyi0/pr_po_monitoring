import streamlit as st
import pandas as pd
import plotly.express as px
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.text import RichText
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.text import Font as DrawingFont, RegularTextRun
from openpyxl.chart.text import Text
from openpyxl.chart.title import Title
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.label import DataLabelList, DataLabel

from utils import MAPPING_SINGKATAN

BULAN_INDO = {
    1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'Mei', 6: 'Juni',
    7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
}

# Label komparasi khusus untuk garis "Harga Perolehan" pada chart.
# Dipakai untuk membedakan garis ini dari garis komparasi Majalah - Incoterm biasa,
# supaya bisa dikecualikan dari perhitungan resume otomatis & tabel histori data.
LABEL_HARGA_PEROLEHAN = "Harga Perolehan"
# Default warna kuning/emas dipilih karena tetap kontras baik di tema terang
# maupun tema gelap (berbeda dengan hitam yang nyaris tak terlihat di dark mode).
WARNA_HARGA_PEROLEHAN_DEFAULT = "#FFC300"


# =============================================================================
# KONFIGURASI SEMUA BAHAN BAKU
# Tambahkan/edit entry di sini untuk menambah bahan baku baru, tanpa file baru.
#
# Field:
#   label                 : Nama tampilan (judul, chart, sheet Excel, dsb)
#   db_value              : Nilai kolom `bahan_baku` di query SQL (biasanya = label)
#   threshold_signifikan  : Batas ($ USD/MT) untuk narasi "signifikan" vs "tidak signifikan"
#   kata_naik / kata_turun: Kata kerja tren, mis. "meningkat"/"menurun" atau "menguat"/"melemah"
#   kalimat_dampak        : Template kalimat dampak industri (opsional, isi None jika tidak ada,
#                           seperti pola Ammonia -> "pupuk fosfat")
#   default_komparasi     : List default Majalah+Incoterm yang tampil pertama kali dibuka,
#                           mis. [{"majalah": "Fertecon Ammonia", "incoterm": "South East Asia FOB"}, ...].
#                           Kosongkan ([]) jika belum ada default khusus -> fallback ke majalah/incoterm
#                           pertama yang tersedia di data (perilaku lama). Jika majalah/incoterm yang
#                           didefinisikan di sini ternyata tidak ada di data, juga fallback diam-diam
#                           ke majalah/incoterm pertama yang tersedia (tanpa menampilkan peringatan).
# =============================================================================
BAHAN_BAKU_CONFIG = {
    "Ammonia": {
        "label": "Ammonia",
        "db_value": "Ammonia",
        "threshold_signifikan": 25.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": "pupuk fosfat",
        "default_komparasi": [
            {"majalah": "Fertecon Ammonia", "incoterm": "South East Asia FOB"},
            {"majalah": "Argus FMB Price Guide", "incoterm": "East Asia CFR (excl Taiwan)"},
        ],
    },
    "DAP": {
        "label": "DAP",
        "db_value": "DAP",
        "threshold_signifikan": 20.0,
        "kata_naik": "menguat",
        "kata_turun": "melemah",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "Argus FMB Price Guide", "incoterm": "China FOB"},
            {"majalah": "Fertecon Phosphate", "incoterm": "FOB China Cash"},
        ],
    },
    "MOP-KCl": {
        "label": "MOP-KCl",
        "db_value": "MOP-KCl",
        "threshold_signifikan": 20.0,
        "kata_naik": "menguat",
        "kata_turun": "melemah",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "MOP Ref. Argus FMB Price Guide (spot)", "incoterm": "SE Asia CFR Spot Std"},
            {"majalah": "CRU", "incoterm": "CFR SEA"},
        ],
    },
    "NH4Cl": {
        "label": "NH4Cl",
        "db_value": "NH4Cl",
        "threshold_signifikan": 10.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "Argus Nitrogen", "incoterm": "CFR Southeast Asia"},
        ],
    },
    "NPK": {
        "label": "NPK",
        "db_value": "NPK",
        "threshold_signifikan": 15.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [],
    },
    "Phosphoric Acid": {
        "label": "Phosphoric Acid",
        "db_value": "Phosphoric Acid",
        "threshold_signifikan": 15.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "Argus FMB Price Guide", "incoterm": "India CFR"},
            {"majalah": "Fertecon Phosphate", "incoterm": "India CFR"},
        ],
    },
    "Phosphate Rock": {
        "label": "Phosphate Rock",
        "db_value": ["phosphate rock", "phos rock", "phosrock"],
        "threshold_signifikan": 15.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "Profercy Phosphate", "incoterm": "Jordan FOB 28-31%"},
            {"majalah": "Profercy Phosphate", "incoterm": "Egypt FOB 30-31%"},
            {"majalah": "Profercy Phosphate", "incoterm": "Jordan FOB 32-34%"},
            {"majalah": "Argus FMB Price Guide", "incoterm": "Jordan FOB 68-70% BPL"},
        ],
    },
    "Sulfur": {
        "label": "Sulfur",
        "db_value": "Sulfur",
        "threshold_signifikan": 10.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "Argus FMB Price Guide Sulphur", "incoterm": "CFR Indonesia Spot"},
            {"majalah": "Argus FMB Price Guide Sulphur", "incoterm": "CFR India Spot"},
            {"majalah": "Fertecon Sulphur", "incoterm": "India CFR"},
        ],
    },
    "Sulfuric Acid": {
        "label": "Sulfuric Acid",
        "db_value": "Sulfuric Acid",
        "threshold_signifikan": 10.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "Majalah ICIS SA pricing", "incoterm": "CFR Indonesia"},
            {"majalah": "Majalah ICIS SA pricing", "incoterm": "CFR S.E. Asia Spot"},
            {"majalah": "Argus FMB Sulphuric Acid", "incoterm": "Spot Price - cfr SEA"},
        ],
    },
    "TSP": {
        "label": "TSP",
        "db_value": "TSP",
        "threshold_signifikan": 10.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [],
    },
    "Urea": {
        "label": "Urea",
        "db_value": "Urea",
        "threshold_signifikan": 15.0,
        "kata_naik": "meningkat",
        "kata_turun": "menurun",
        "kalimat_dampak": None,
        "default_komparasi": [],
    },
    "ZA": {
        "label": "ZA",
        "db_value": "ZA",
        "threshold_signifikan": 10.0,
        "kata_naik": "menguat",
        "kata_turun": "melemah",
        "kalimat_dampak": None,
        "default_komparasi": [
            {"majalah": "Fertecon Nitrates", "incoterm": "SE Asia CFR Caprolactam"},
            {"majalah": "Argus FMB Price Guide Nitrogen", "incoterm": "(NH4)2SO4 SE Asia CFR"},
        ],
    },
}


def get_daftar_bahan_baku():
    return list(BAHAN_BAKU_CONFIG.keys())


def get_config(bahan_baku_key):
    if bahan_baku_key not in BAHAN_BAKU_CONFIG:
        raise KeyError(f"Konfigurasi untuk bahan baku '{bahan_baku_key}' belum didefinisikan di BAHAN_BAKU_CONFIG")
    return BAHAN_BAKU_CONFIG[bahan_baku_key]


# =============================================================================
# HARGA PEROLEHAN: query data dari tabel terpisah `harga_perolehan_bahan_baku`
# =============================================================================
def _load_harga_perolehan(load_data, db_value, start_date, end_date):
    """
    Mengambil data Harga Perolehan (tabel harga_perolehan_bahan_baku) untuk
    bahan_baku terpilih, pada rentang tanggal yang sama dengan filter chart.
    Mengembalikan DataFrame kosong jika tidak ada data (mis. tabel belum ada,
    atau bahan baku tersebut memang tidak punya kolom Harga Perolehan di sumbernya).
    """
    if isinstance(db_value, (list, tuple)):
        alias_list = "', '".join([a.lower().strip() for a in db_value])
        where_clause = f"lower(trim(bahan_baku)) IN ('{alias_list}')"
    else:
        where_clause = f"bahan_baku = '{db_value}'"

    query = f"""
        SELECT tanggal_terbit, harga_perolehan
        FROM harga_perolehan_bahan_baku
        WHERE {where_clause}
          AND tanggal_terbit >= '{start_date}' AND tanggal_terbit <= '{end_date}'
        ORDER BY tanggal_terbit ASC
    """
    try:
        df_hp = load_data(query)
    except Exception:
        # Tabel mungkin belum ada di database (belum pernah dijalankan ETL Harga Perolehan)
        return pd.DataFrame(columns=['tanggal_terbit', 'harga_perolehan'])

    return df_hp


# =============================================================================
# RESUME OTOMATIS GENERIK (menggantikan hitung_resume_ammonia, hitung_resume_dap, dst)
# =============================================================================
def _get_nama_minggu(dt):
    if dt.day <= 7:
        return f"awal {BULAN_INDO[dt.month]} {dt.year}"
    elif dt.day <= 14:
        return f"minggu kedua {BULAN_INDO[dt.month]} {dt.year}"
    elif dt.day <= 21:
        return f"minggu ketiga {BULAN_INDO[dt.month]} {dt.year}"
    else:
        return f"akhir {BULAN_INDO[dt.month]} {dt.year}"


def hitung_resume_generik(df_plot, y_col, config):
    """
    Fungsi resume otomatis generik untuk semua bahan baku.
    Perbedaan narasi antar bahan baku (threshold, kata tren, kalimat dampak)
    di-drive lewat `config` (lihat BAHAN_BAKU_CONFIG di atas), bukan lewat
    duplikasi fungsi per bahan baku.

    Catatan: `df_plot` yang diterima di sini HARUS sudah tidak mengandung baris
    Harga Perolehan (label_komparasi == LABEL_HARGA_PEROLEHAN), karena resume
    ini murni bicara soal komparasi Majalah - Incoterm.
    """
    label_bb = config["label"]
    threshold_signifikan = config.get("threshold_signifikan", 25.0)
    kata_naik = config.get("kata_naik", "meningkat")
    kata_turun = config.get("kata_turun", "menurun")
    kalimat_dampak = config.get("kalimat_dampak")

    if df_plot.empty:
        return ["Data tidak tersedia."]

    tgl_T0 = pd.Timestamp(df_plot['tanggal_terbit'].max())
    batas_1_bulan = tgl_T0 - pd.DateOffset(months=1)
    batas_2_bulan = tgl_T0 - pd.DateOffset(months=2)

    df_T0 = df_plot.sort_values('tanggal_terbit').drop_duplicates(subset=['label_komparasi'], keep='last')
    harga_T0 = df_T0[y_col].mean()

    df_T1_range = df_plot[(df_plot['tanggal_terbit'] >= batas_1_bulan) & (df_plot['tanggal_terbit'] < tgl_T0)]
    if not df_T1_range.empty:
        df_T1 = df_T1_range.sort_values('tanggal_terbit').drop_duplicates(subset=['label_komparasi'], keep='last')
        harga_T1 = df_T1[y_col].mean()
        tgl_T1 = pd.Timestamp(df_T1['tanggal_terbit'].max())
    else:
        harga_T1, tgl_T1 = harga_T0, tgl_T0

    df_T2_range = df_plot[(df_plot['tanggal_terbit'] >= batas_2_bulan) & (df_plot['tanggal_terbit'] < batas_1_bulan)]
    if not df_T2_range.empty:
        df_T2 = df_T2_range.sort_values('tanggal_terbit').drop_duplicates(subset=['label_komparasi'], keep='first')
        harga_T2 = df_T2[y_col].mean()
        tgl_T2 = pd.Timestamp(df_T2['tanggal_terbit'].min())
    else:
        df_T2 = df_plot.sort_values('tanggal_terbit').drop_duplicates(subset=['label_komparasi'], keep='first')
        harga_T2 = df_T2[y_col].mean()
        tgl_T2 = pd.Timestamp(df_T2['tanggal_terbit'].min())

    delta_recent = harga_T0 - harga_T1
    delta_past = harga_T1 - harga_T2

    if delta_recent < -2.0:
        tren_sekarang = f"menunjukkan tren {kata_turun}"
        signifikansi = " namun tidak signifikan" if abs(delta_recent) < threshold_signifikan else " yang cukup signifikan"
        konteks_historis = f" setelah mengalami kenaikan sepanjang {BULAN_INDO[tgl_T1.month]} {tgl_T1.year}" if delta_past > 2.0 else ""
    elif delta_recent > 2.0:
        tren_sekarang = f"menunjukkan tren {kata_naik}"
        signifikansi = " namun tidak signifikan" if abs(delta_recent) < threshold_signifikan else " yang cukup signifikan"
        konteks_historis = f" setelah mengalami penurunan sepanjang {BULAN_INDO[tgl_T1.month]} {tgl_T1.year}" if delta_past < -2.0 else ""
    else:
        tren_sekarang = "terpantau stabil"
        signifikansi, konteks_historis = "", ""

    poin_1 = f"Secara keseluruhan, harga {label_bb} {tren_sekarang}{signifikansi} pada {_get_nama_minggu(tgl_T0.date())}{konteks_historis}."

    is_turun = tren_sekarang == f"menunjukkan tren {kata_turun}"
    is_naik = tren_sekarang == f"menunjukkan tren {kata_naik}"

    if is_turun:
        if kalimat_dampak:
            if harga_T0 > harga_T2:
                poin_2 = (f"Meskipun menunjukkan tren {kata_turun}, harga {label_bb} masih bertahan pada level yang "
                          f"tinggi dibandingkan awal {BULAN_INDO[tgl_T2.month]} {tgl_T2.year}, sehingga tetap memberikan "
                          f"tekanan terhadap biaya produksi {kalimat_dampak}.")
            else:
                poin_2 = (f"Penurunan harga ini sedikit memberikan kelonggaran terhadap tekanan biaya produksi "
                          f"{kalimat_dampak} jika dibandingkan rata-rata periode {BULAN_INDO[tgl_T2.month]}.")
        else:
            if harga_T0 > harga_T2:
                poin_2 = (f"Meskipun demikian, posisi harga rata-rata saat ini masih bertahan pada level yang cenderung "
                          f"lebih tinggi jika dibandingkan dengan periode awal {BULAN_INDO[tgl_T2.month]} {tgl_T2.year}.")
            elif harga_T0 < harga_T2:
                poin_2 = (f"Penurunan ini membawa rata-rata harga pasar bergerak ke level yang lebih rendah terpaut "
                          f"USD {abs(harga_T0 - harga_T2):.2f}/MT dari posisi baseline awal {BULAN_INDO[tgl_T2.month]}.")
            else:
                poin_2 = f"Harga bergerak konstan dan stabil mereplikasi pergerakan harga pada periode awal {BULAN_INDO[tgl_T2.month]}."

    elif is_naik:
        if kalimat_dampak:
            poin_2 = (f"Peningkatan harga {label_bb} ini semakin memberikan tekanan berat terhadap struktur biaya "
                      f"produksi {kalimat_dampak} karena posisi harga bergerak menjauh dari rata-rata baseline "
                      f"USD {harga_T2:.2f}/MT.")
        else:
            poin_2 = (f"Peningkatan ini membawa rata-rata harga pasar bergerak menjauh dari posisi baseline awal "
                      f"{BULAN_INDO[tgl_T2.month]} sebesar USD {harga_T2:.2f}/MT.")
    else:
        poin_2 = (f"Harga {label_bb} terpantau bertahan stabil pada level konstan jika dibandingkan dengan rata-rata "
                  f"periode awal {BULAN_INDO[tgl_T2.month]} {tgl_T2.year} yang berada di rata-rata USD {harga_T2:.2f}/MT.")

    list_resume = [poin_1, poin_2]

    for label in df_plot['label_komparasi'].unique():
        df_ref = df_plot[df_plot['label_komparasi'] == label]
        max_tgl_ref = pd.Timestamp(df_ref['tanggal_terbit'].max())
        if (tgl_T0 - max_tgl_ref).days > 14:
            tgl_str = f"{max_tgl_ref.day:02d} {BULAN_INDO[max_tgl_ref.month]} {max_tgl_ref.year}"
            list_resume.append(f"Untuk referensi {label}, harga terakhir dirilis pada {tgl_str}.")

    return list_resume


# =============================================================================
# EXCEL EXPORT (identik untuk semua bahan baku, cuma judul & sheet name beda)
# =============================================================================
def generate_excel_export(df_plot, df_pivot, kolom_tanggal, y_col, y_label, jenis_harga, warna_map, list_resume, label_bb):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Komparasi Harga {label_bb}"[:31]  # Excel sheet name max 31 char
    ws_data = wb.create_sheet("_DataChart")

    df_chart = df_plot.pivot_table(
        index='tanggal_terbit', columns='label_komparasi', values=y_col, aggfunc='mean'
    ).sort_index()

    df_chart_reset = df_chart.reset_index()
    df_chart_reset['tanggal_terbit'] = df_chart_reset['tanggal_terbit'].dt.strftime('%d %b %Y')

    headers = ['Tanggal Terbit'] + list(df_chart.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws_data.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(df_chart_reset.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            # Format kolom data nilai (bukan tanggal) agar muncul 2 desimal di chart
            if col_idx > 1 and isinstance(value, (int, float)) and not pd.isna(value):
                cell.number_format = '#,##0.00'

    n_rows = len(df_chart_reset)
    n_cols = len(headers)

    arial_font = DrawingFont(typeface='Arial')
    cp_arial = CharacterProperties(latin=arial_font)
    cp_arial_bold = CharacterProperties(latin=arial_font, b=True)
    cp_arial_sz700 = CharacterProperties(sz=700, latin=arial_font)

    def create_formatted_title(text_val, is_bold=True):
        cp = cp_arial_bold if is_bold else cp_arial
        run = RegularTextRun(t=text_val, rPr=cp)
        p = Paragraph(pPr=ParagraphProperties(defRPr=cp), r=[run])
        return Title(tx=Text(rich=RichText(p=[p])))

    is_single_series = (n_cols - 1) == 1
    if is_single_series:
        dummy_col_idx = n_cols + 1
        ws_data.cell(row=1, column=dummy_col_idx, value=None)
        n_cols = dummy_col_idx

    chart = LineChart()
    chart.title = create_formatted_title(f"Komparasi Tren Harga {label_bb} ({jenis_harga})", is_bold=True)
    chart.height = 14
    chart.width = 32
    chart.style = None
    chart.title.overlay = False

    data_ref = Reference(ws_data, min_col=2, max_col=n_cols, min_row=1, max_row=n_rows + 1)
    cats_ref = Reference(ws_data, min_col=1, max_col=1, min_row=2, max_row=n_rows + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    label_columns = list(df_chart.columns)
    for series, label in zip(chart.series, label_columns):
        series.marker.symbol = "none"
        series.smooth = False
        hex_color = warna_map.get(label, "#1f77b4").lstrip('#').upper()
        series.graphicalProperties.line.width = 18000
        series.graphicalProperties.line.solidFill = hex_color

        # --- [MODIFIKASI 2 MULAI]: MEMUNCULKAN ANGKA DI TITIK TERAKHIR ---
        # 1. Cari nama index dari baris data valid terakhir (bukan NaN) di kolom terkait
        last_valid_idx_label = df_chart[label].last_valid_index()
        
        if last_valid_idx_label is not None:
            last_idx = df_chart.index.get_loc(last_valid_idx_label)

            # 1. Inisialisasi daftar label untuk series ini
            series.dLbls = DataLabelList()
            
            # 2. MATIKAN SEMUA pengaturan label global agar tidak menumpuk di semua titik
            series.dLbls.showVal = False
            series.dLbls.showCatName = False
            series.dLbls.showSerName = False
            series.dLbls.showPercent = False
            series.dLbls.showLegendKey = False
            series.dLbls.showBubbleSize = False
            
            # 3. Buat pengaturan label khusus HANYA untuk titik terakhir (last_idx)
            dl = DataLabel(idx=last_idx)
            dl.showVal = True         # Hanya nyalakan angkanya
            dl.showCatName = False    # Pastikan tanggal tidak ikut muncul
            dl.showSerName = False    # Pastikan nama majalah/incoterm tidak ikut muncul
            
            # 4. Sematkan label khusus tersebut ke series
            series.dLbls.dLbl.append(dl)
        # --- [MODIFIKASI 2 SELESAI] ---

    if is_single_series:
        dummy_series = chart.series[-1]
        dummy_series.graphicalProperties.line.noFill = True
        dummy_series.marker.symbol = "none"
        dummy_idx = len(chart.series) - 1
        if chart.legend.legendEntry is None:
            chart.legend.legendEntry = []
        chart.legend.legendEntry.append(LegendEntry(idx=dummy_idx, delete=True))

    chart.y_axis.title = create_formatted_title(y_label, is_bold=True)
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_arial), endParaRPr=cp_arial)])
    chart.y_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines.graphicalProperties = GraphicalProperties()
    chart.y_axis.majorGridlines.graphicalProperties.line = LineProperties(solidFill="E0E0E0", w=9525)
    chart.y_axis.delete = False

    chart.x_axis.title = create_formatted_title("Tanggal Publikasi", is_bold=True)
    chart.x_axis.txPr = RichText(
        bodyPr=RichTextProperties(rot=-5400000, vert="horz"),
        p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_arial_sz700), endParaRPr=cp_arial_sz700)]
    )
    chart.x_axis.delete = False
    chart.x_axis.majorGridlines = None
    chart.x_axis.tickLblSkip = 1
    chart.x_axis.tickMarkSkip = 1

    chart.legend.position = 'b'
    chart.legend.overlay = False
    chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp_arial), endParaRPr=cp_arial)])

    chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.18, h=0.64, w=0.90, xMode="edge", yMode="edge"))
    ws.add_chart(chart, "A1")

    # ========================== STYLING TABEL ==========================
    HEADER_BLUE = "BDD7EE"
    thin = Side(style='thin', color='000000')
    TABLE_START_ROW = 34
    n_date_cols = len(kolom_tanggal)
    last_col = 1 + n_date_cols

    ws.cell(row=TABLE_START_ROW, column=1, value="Detail Histori Data (3 Periode Terakhir)").font = Font(bold=True, size=13)

    header_row1 = TABLE_START_ROW + 1
    header_row2 = header_row1 + 1
    first_data_row = header_row2 + 1
    last_data_row = first_data_row + len(df_pivot) - 1

    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=header_row1, column=col_idx)
        cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=header_row1, column=1, value="Referensi")
    ws.cell(row=header_row1, column=2, value="Harga USD/MT")

    ws.cell(row=header_row2, column=1).fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    for col_idx, tgl_label in enumerate(kolom_tanggal, start=2):
        cell = ws.cell(row=header_row2, column=col_idx, value=tgl_label)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, (index_label, row) in enumerate(df_pivot.iterrows()):
        r = first_data_row + row_offset
        cell = ws.cell(row=r, column=1, value=index_label)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx, col in enumerate(df_pivot.columns, start=2):
            val = row[col]
            val_str = "" if pd.isna(val) else str(val)
            c = ws.cell(row=r, column=col_idx, value=val_str)
            c.alignment = Alignment(horizontal="center", vertical="center")

    border_thin_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(header_row1, last_data_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c).border = border_thin_all

    ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)
    ws.merge_cells(start_row=header_row1, start_column=2, end_row=header_row1, end_column=last_col)

    # ========================== RESUME ==========================
    resume_title_row = last_data_row + 2
    ws.cell(row=resume_title_row, column=1, value="Resume :").font = Font(bold=True, size=11, italic=True, name='Arial')

    for idx, poin in enumerate(list_resume, start=1):
        current_resume_row = resume_title_row + idx
        cell = ws.cell(row=current_resume_row, column=1, value=f"•  {poin}")
        cell.font = Font(size=11, name='Arial')
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        ws.merge_cells(start_row=current_resume_row, start_column=1, end_row=current_resume_row, end_column=last_col)
        jumlah_baris = (len(poin) // 90) + 1
        ws.row_dimensions[current_resume_row].height = 16 * jumlah_baris

    ws.column_dimensions['A'].width = 26
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    ws.row_dimensions[header_row1].height = 22
    ws.row_dimensions[header_row2].height = 20

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    is_bold = cell.font.bold if cell.font else False
                    is_italic = cell.font.italic if cell.font else False
                    cell.font = Font(name='Arial', size=11, bold=is_bold, italic=is_italic)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def variasikan_warna(hex_color, index, total):
    if total <= 1:
        return hex_color
    hex_color = hex_color.lstrip('#')
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    factor = 0.6 + (0.7 * index / max(total - 1, 1))
    r = min(255, int(r * factor))
    g = min(255, int(g * factor))
    b = min(255, int(b * factor))
    return f"#{r:02x}{g:02x}{b:02x}"


# =============================================================================
# RENDER UTAMA (dipanggil sekali per bahan baku terpilih)
# =============================================================================
def render(load_data, global_context):
    st.markdown("### :material/science: Analisis Tren Komparasi Harga Pasar Bahan Baku")

    daftar_bb = get_daftar_bahan_baku()
    bahan_baku_pilihan = st.selectbox("Pilih Bahan Baku", daftar_bb, key="pilihan_bahan_baku")

    config = get_config(bahan_baku_pilihan)
    label_bb = config["label"]
    db_value = config["db_value"]

    # Suffix unik untuk session_state key, supaya filter tiap bahan baku tidak tabrakan
    suffix = bahan_baku_pilihan.lower().replace(" ", "_")

    st.markdown(f"#### :material/science: {label_bb}")

    from config_db import get_setting

    bahan_baku_date_str = get_setting("DATA_UPDATE_BAHAN_BAKU", "2026-03-31")
    try:
        tgl_update_bb = datetime.strptime(bahan_baku_date_str, "%Y-%m-%d").date()
    except Exception:
        tgl_update_bb = datetime(2026, 3, 31).date()

    tgl_update_str = f"{tgl_update_bb.day:02d} {BULAN_INDO[tgl_update_bb.month]} {tgl_update_bb.year}"

    st.markdown(
        f"<p style='font-size:14px; opacity:0.65; margin-top:-6px; margin-bottom:16px;'>"
        f"Data terakhir diperbarui pada <b>{tgl_update_str}</b>"
        f"</p>",
        unsafe_allow_html=True
    )

    if isinstance(db_value, (list, tuple)):
        alias_list = "', '".join([a.lower().strip() for a in db_value])
        where_clause = f"lower(trim(bahan_baku)) IN ('{alias_list}')"
    else:
        where_clause = f"bahan_baku = '{db_value}'"

    query = f"""
        SELECT tanggal_terbit, nama_majalah, incoterm, harga_min, harga_max 
        FROM master_harga_bahan_baku 
        WHERE {where_clause}
        ORDER BY tanggal_terbit ASC
    """
    df = load_data(query)

    if df.empty:
        st.warning(f"Data harga {label_bb} belum tersedia di database.")
        return

    list_majalah = df['nama_majalah'].unique()
    min_date = df['tanggal_terbit'].min()
    max_date = df['tanggal_terbit'].max()

    today = datetime.now()
    default_start_date = pd.Timestamp(year=today.year, month=today.month, day=1) - pd.DateOffset(months=14)
    default_start_date = default_start_date.date()
    calendar_min_date = min(min_date, default_start_date)

    if default_start_date > max_date or default_start_date < min_date:
        default_start_date = min_date

    def _save_to_permanent(widget_key, permanent_key):
        st.session_state[permanent_key] = st.session_state[widget_key]

    with st.expander(":material/settings: Filter Komparasi Harga Pasar", expanded=True):
        col_mulai, col_sampai, col_metode, col_jml = st.columns(4)
        with col_mulai:
            start_date = st.date_input(
                "Mulai dari tanggal",
                value=st.session_state.get(f"_perm_start_date_{suffix}", default_start_date),
                min_value=calendar_min_date, max_value=max_date,
                key=f"start_date_{suffix}",
                on_change=_save_to_permanent,
                args=(f"start_date_{suffix}", f"_perm_start_date_{suffix}")
            )
        with col_sampai:
            end_date = st.date_input(
                "Sampai tanggal",
                value=st.session_state.get(f"_perm_end_date_{suffix}", max_date),
                min_value=calendar_min_date, max_value=max_date,
                key=f"end_date_{suffix}",
                on_change=_save_to_permanent,
                args=(f"end_date_{suffix}", f"_perm_end_date_{suffix}")
            )
        with col_metode:
            jenis_harga_options = ["AVERAGE", "MIN", "MAX"]
            jenis_harga_default = st.session_state.get(f"_perm_jenis_harga_{suffix}", "AVERAGE")
            jenis_harga = st.selectbox(
                "Jenis Harga", jenis_harga_options,
                index=jenis_harga_options.index(jenis_harga_default) if jenis_harga_default in jenis_harga_options else 0,
                help="Pilih nilai harga yang ingin diplot pada grafik",
                key=f"jenis_harga_{suffix}",
                on_change=_save_to_permanent,
                args=(f"jenis_harga_{suffix}", f"_perm_jenis_harga_{suffix}")
            )
        with col_jml:
            # Default jumlah komparasi mengikuti panjang `default_komparasi` di config
            # (jika ada dan belum ada pilihan tersimpan dari interaksi sebelumnya).
            default_komparasi_list = config.get("default_komparasi", [])
            default_jml_komparasi = len(default_komparasi_list) if default_komparasi_list else 2
            jml_komparasi = st.number_input(
                "Jumlah Komparasi", min_value=1, max_value=5,
                value=st.session_state.get(f"_perm_jml_komparasi_{suffix}", default_jml_komparasi),
                key=f"jml_komparasi_{suffix}",
                on_change=_save_to_permanent,
                args=(f"jml_komparasi_{suffix}", f"_perm_jml_komparasi_{suffix}")
            )

        st.markdown("<hr style='margin: 10px 0; border-color: rgba(255,255,255,0.1);'>", unsafe_allow_html=True)

        # Toggle untuk menampilkan/menyembunyikan garis Harga Perolehan di chart,
        # beserta color picker khusus untuk garis tersebut. Defaultnya aktif;
        # kalau datanya memang tidak ada, toggle ini otomatis tidak berpengaruh
        # karena tidak ada apapun yang ditambahkan ke chart.
        col_hp_toggle, col_hp_warna = st.columns([3, 1])
        with col_hp_toggle:
            tampilkan_harga_perolehan = st.checkbox(
                "Tampilkan garis Harga Perolehan pada chart",
                value=st.session_state.get(f"_perm_tampilkan_hp_{suffix}", True),
                key=f"tampilkan_hp_{suffix}",
                on_change=_save_to_permanent,
                args=(f"tampilkan_hp_{suffix}", f"_perm_tampilkan_hp_{suffix}")
            )
        with col_hp_warna:
            warna_harga_perolehan = st.color_picker(
                "Warna",
                value=st.session_state.get(f"_perm_warna_hp_{suffix}", WARNA_HARGA_PEROLEHAN_DEFAULT),
                key=f"warna_hp_{suffix}",
                on_change=_save_to_permanent,
                args=(f"warna_hp_{suffix}", f"_perm_warna_hp_{suffix}")
            )

        st.markdown("<hr style='margin: 10px 0; border-color: rgba(255,255,255,0.1);'>", unsafe_allow_html=True)

        komparasi_data = []
        warna_map = {}
        default_colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd"]

        for i in range(int(jml_komparasi)):
            # Default majalah & incoterm untuk index ke-i, diambil dari `default_komparasi`
            # di config (kalau ada entry untuk index ini). Jika majalah/incoterm yang
            # dikonfigurasi ternyata tidak ada di data (mis. salah ketik / data belum masuk),
            # fallback diam-diam ke majalah/incoterm pertama yang tersedia.
            default_dari_config = default_komparasi_list[i] if i < len(default_komparasi_list) else None

            c1, c2, c3 = st.columns([3, 3, 1])
            with c1:
                perm_key_majalah = f"_perm_majalah_{suffix}_{i}"
                if perm_key_majalah in st.session_state:
                    default_majalah = st.session_state[perm_key_majalah]
                elif default_dari_config and default_dari_config.get("majalah") in list_majalah:
                    default_majalah = default_dari_config["majalah"]
                else:
                    default_majalah = list_majalah[i] if i < len(list_majalah) else list_majalah[0]
                majalah_index = list(list_majalah).index(default_majalah) if default_majalah in list_majalah else 0
                majalah_pilihan = st.selectbox(
                    f"Majalah ke-{i+1}", list_majalah,
                    index=majalah_index,
                    key=f"majalah_{suffix}_{i}",
                    on_change=_save_to_permanent,
                    args=(f"majalah_{suffix}_{i}", perm_key_majalah)
                )
            with c2:
                list_incoterm = df[df['nama_majalah'] == majalah_pilihan]['incoterm'].unique()
                perm_key_incoterm = f"_perm_incoterm_{suffix}_{i}"
                if perm_key_incoterm in st.session_state:
                    default_incoterm = st.session_state[perm_key_incoterm]
                elif (default_dari_config and majalah_pilihan == default_dari_config.get("majalah")
                      and default_dari_config.get("incoterm") in list_incoterm):
                    default_incoterm = default_dari_config["incoterm"]
                else:
                    default_incoterm = list_incoterm[0] if len(list_incoterm) > 0 else None
                incoterm_index = list(list_incoterm).index(default_incoterm) if default_incoterm in list_incoterm else 0
                incoterm_pilihan = st.selectbox(
                    f"Metode Incoterm ke-{i+1}", list_incoterm,
                    index=incoterm_index if len(list_incoterm) > 0 else None,
                    key=f"incoterm_{suffix}_{i}",
                    on_change=_save_to_permanent,
                    args=(f"incoterm_{suffix}_{i}", perm_key_incoterm)
                )
            with c3:
                perm_key_warna = f"_perm_warna_{suffix}_{i}"
                default_warna = st.session_state.get(perm_key_warna, default_colors[i % len(default_colors)])
                warna_pilihan = st.color_picker(
                    "Warna", default_warna,
                    key=f"color_{suffix}_{i}",
                    on_change=_save_to_permanent,
                    args=(f"color_{suffix}_{i}", perm_key_warna)
                )

            if incoterm_pilihan:
                komparasi_data.append({
                    "majalah": majalah_pilihan,
                    "incoterms": [incoterm_pilihan],
                    "warna_dasar": warna_pilihan
                })

        for item in komparasi_data:
            for idx, incoterm in enumerate(item["incoterms"]):
                label_asli = f"{item['majalah']} - {incoterm}"
                label_singkat = MAPPING_SINGKATAN.get(label_asli, label_asli)
                warna_final = variasikan_warna(item["warna_dasar"], idx, len(item["incoterms"]))
                warna_map[label_singkat] = warna_final

    col_btn, _ = st.columns([1, 4])
    with col_btn:
        if st.button(":material/refresh: Refresh Data", use_container_width=True, key=f"refresh_{suffix}"):
            st.cache_data.clear()
            st.rerun()

    if start_date <= end_date and komparasi_data:
        df_plot = pd.DataFrame()

        for item in komparasi_data:
            majalah = item["majalah"]
            incoterms = item["incoterms"]
            temp_df = df[(df['nama_majalah'] == majalah) & (df['incoterm'].isin(incoterms)) &
                         (df['tanggal_terbit'] >= start_date) & (df['tanggal_terbit'] <= end_date)].copy()
            if not temp_df.empty:
                temp_df['label_komparasi'] = temp_df['nama_majalah'] + ' - ' + temp_df['incoterm']
                temp_df['label_komparasi'] = temp_df['label_komparasi'].apply(
                    lambda x: MAPPING_SINGKATAN.get(x, x)
                )
                df_plot = pd.concat([df_plot, temp_df], ignore_index=True)

        if not df_plot.empty:
            df_plot['harga_avg'] = (df_plot['harga_min'] + df_plot['harga_max']) / 2
            df_plot['tanggal_terbit'] = pd.to_datetime(df_plot['tanggal_terbit'])
            df_plot = df_plot.sort_values('tanggal_terbit')

            # Simpan salinan df_plot SEBELUM Harga Perolehan ditambahkan, untuk dipakai
            # oleh tabel "Detail Histori Data", resume otomatis, dan Excel export.
            # Dengan begitu ketiganya tetap murni bicara soal komparasi Majalah - Incoterm.
            df_plot_komparasi = df_plot.copy()

            if jenis_harga == "MIN":
                y_col, y_label = 'harga_min', 'Harga Minimum (USD/MT)'
            elif jenis_harga == "MAX":
                y_col, y_label = 'harga_max', 'Harga Maksimum (USD/MT)'
            else:
                y_col, y_label = 'harga_avg', 'Harga Rata-rata (USD/MT)'

            # ============= Tambahkan garis Harga Perolehan (jika ada & diaktifkan) =============
            df_hp = pd.DataFrame()
            if tampilkan_harga_perolehan:
                df_hp = _load_harga_perolehan(load_data, db_value, start_date, end_date)

            df_plot_chart = df_plot.copy()
            if not df_hp.empty:
                df_hp = df_hp.copy()
                df_hp['tanggal_terbit'] = pd.to_datetime(df_hp['tanggal_terbit'])
                df_hp['label_komparasi'] = LABEL_HARGA_PEROLEHAN
                # Kolom Harga Perolehan diisi ke y_col yang sedang aktif (MIN/MAX/AVERAGE),
                # karena Harga Perolehan sendiri hanya berupa satu nilai per tanggal
                # (bukan rentang min-max), jadi nilainya sama untuk ketiga jenis harga.
                df_hp[y_col] = df_hp['harga_perolehan']
                df_hp_for_plot = df_hp[['tanggal_terbit', 'label_komparasi', y_col]]
                df_plot_chart = pd.concat([df_plot_chart, df_hp_for_plot], ignore_index=True)
                warna_map[LABEL_HARGA_PEROLEHAN] = warna_harga_perolehan

            tanggal_unik = df_plot_chart['tanggal_terbit'].unique()

            fig = px.line(
                df_plot_chart, x='tanggal_terbit', y=y_col, color='label_komparasi',
                color_discrete_map=warna_map,
                title=f"Komparasi Tren Harga {label_bb} ({jenis_harga})",
                labels={y_col: y_label, 'tanggal_terbit': 'Tanggal Publikasi', 'label_komparasi': 'Majalah & Incoterm'}
            )

            # Garis Harga Perolehan dibedakan secara visual (putus-putus) supaya
            # tidak tertukar dengan garis komparasi Majalah - Incoterm biasa.
            if not df_hp.empty:
                fig.for_each_trace(
                    lambda tr: tr.update(line=dict(dash="dash", width=3)) if tr.name == LABEL_HARGA_PEROLEHAN else ()
                )

            # Label angka pada titik data TERAKHIR tiap garis (komparasi maupun Harga
            # Perolehan), warnanya mengikuti warna garis masing-masing, supaya nilai
            # terkini langsung terbaca tanpa perlu hover.
            for label, df_label in df_plot_chart.groupby('label_komparasi'):
                df_label_sorted = df_label.sort_values('tanggal_terbit')
                titik_terakhir = df_label_sorted.iloc[-1]
                warna_label = warna_map.get(label, "#1f77b4")
                fig.add_annotation(
                    x=titik_terakhir['tanggal_terbit'],
                    y=titik_terakhir[y_col],
                    text=f"<b>{titik_terakhir[y_col]:.2f}</b>",
                    showarrow=False,
                    xanchor="left",
                    yanchor="middle",
                    xshift=8,
                    font=dict(color=warna_label, size=12),
                    bgcolor="rgba(255,255,255,0.75)",
                )

            fig.update_layout(
                hovermode="x unified",
                legend=dict(orientation="v", yanchor="top", y=-0.6, xanchor="left", x=0),
                margin=dict(b=300, t=80, l=60, r=90),
                height=600
            )

            fig.update_xaxes(
                tickangle=-90, type='date', tickmode='array', tickvals=tanggal_unik,
                tickformat="%d %b %Y", title=dict(text="Tanggal Publikasi", standoff=40)
            )

            fig.update_yaxes(dtick=50)

            st.plotly_chart(fig, use_container_width=True)

            st.markdown("#### :material/table_chart: Detail Histori Data (3 Periode Terakhir)")

            df_display = df_plot_komparasi.copy()
            df_display['harga_range'] = df_display['harga_min'].apply(lambda x: f"{x:.2f}") + ' - ' + df_display['harga_max'].apply(lambda x: f"{x:.2f}")

            df_pivot = df_display.pivot_table(
                index='label_komparasi', columns='tanggal_terbit', values='harga_range',
                aggfunc=lambda x: ' '.join(x)
            )

            df_pivot = df_pivot.sort_index(axis=1, ascending=False)
            df_pivot = df_pivot.iloc[:, :3]

            kolom_tanggal = [f"{d.day:02d} {BULAN_INDO[d.month]} {d.year}" for d in df_pivot.columns]
            jml_kolom = len(kolom_tanggal)

            thead = f'''
<thead>
    <tr>
        <th rowspan="2" style="vertical-align: middle; text-align: left !important;">Referensi</th>
        <th colspan="{jml_kolom}">Harga USD/MT</th>
    </tr>
    <tr>
'''
            for tgl in kolom_tanggal:
                thead += f"<th>{tgl}</th>"
            thead += "</tr>\n</thead>"

            tbody = "<tbody>\n"
            for index, row in df_pivot.iterrows():
                tbody += f"<tr>\n<td style='text-align: left !important;'>{index}</td>\n"
                for col in df_pivot.columns:
                    val = row[col]
                    val_str = "" if pd.isna(val) else str(val)
                    tbody += f"<td>{val_str}</td>\n"
                tbody += "</tr>\n"
            tbody += "</tbody>"

            html_table = f"<table>\n{thead}\n{tbody}\n</table>"

            styled_html = f"""
<style>
.custom-table-container {{
    width: 100%;
    overflow-x: auto;
    margin-bottom: 2rem;
}}
.custom-table-container table {{
    width: 100%;
    border-collapse: collapse;
    font-family: "Source Sans Pro", sans-serif;
    font-size: 14px;
    color: var(--text-color);
}}
.custom-table-container th, .custom-table-container td {{
    text-align: center !important;
    padding: 10px !important;
    border: 1px solid rgba(128, 128, 128, 0.2);
}}
.custom-table-container th {{
    background-color: rgba(128, 128, 128, 0.1);
    font-weight: 600;
}}
</style>
<div class="custom-table-container">
    {html_table}
</div>
"""
            st.markdown(styled_html, unsafe_allow_html=True)

            list_resume_otomatis = hitung_resume_generik(df_plot_komparasi, y_col, config)

            st.markdown("##### *Resume :*")
            for poin in list_resume_otomatis:
                st.markdown(f"- {poin}")
            st.markdown("<br>", unsafe_allow_html=True)

            excel_buffer = generate_excel_export(
                df_plot=df_plot_komparasi, df_pivot=df_pivot, kolom_tanggal=kolom_tanggal,
                y_col=y_col, y_label=y_label, jenis_harga=jenis_harga, warna_map=warna_map,
                list_resume=list_resume_otomatis, label_bb=label_bb
            )

            st.markdown("""
                <style>
                div[data-testid="stDownloadButton"] button {
                    background-color: #FF4B4B;
                    color: white;
                    border: none;
                }
                div[data-testid="stDownloadButton"] button:hover {
                    background-color: #E54444;
                    color: white;
                    border: none;
                }
                div[data-testid="stDownloadButton"] button:active {
                    background-color: #CE3D3D;
                    color: white;
                }
                </style>
            """, unsafe_allow_html=True)

            st.download_button(
                label=":material/download: Download Excel (Chart + Tabel + Resume)",
                data=excel_buffer,
                file_name=f"komparasi_harga_{label_bb}_{jenis_harga}_{start_date}_{end_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            st.info("Tidak ada data yang tersedia untuk kombinasi filter yang dipilih pada rentang waktu tersebut.")
    else:
        if start_date > end_date:
            st.error("❌ 'Mulai dari tanggal' tidak boleh lebih besar dari 'Sampai tanggal'.")
        else:
            st.info("Silakan tentukan minimal 1 metode Incoterm.")
