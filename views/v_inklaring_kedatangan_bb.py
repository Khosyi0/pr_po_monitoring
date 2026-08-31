"""
v_inklaring_kedatangan_bb.py - Halaman Rencana Kedatangan Bahan Baku
Berada di bawah menu "Inklaring Barang Impor", setelah "Analisis Waktu Proses Inklaring".

3 Tab:
  1. Upload File Excel      -> upload file "Rencana Kedatangan Kapal Bahan Baku"
  2. Preview Data Excel     -> preview seluruh isi sheet yang diupload
  3. Preview Hasil Konversi -> filter bulan + kurs manual, hasil siap download
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
import io
from datetime import datetime

# =============================================================================
# KONSTANTA & HELPER PARSING
# =============================================================================

BULAN_MAP = {
    'jan': 1, 'januari': 1,
    'feb': 2, 'februari': 2,
    'mar': 3, 'maret': 3,
    'apr': 4, 'april': 4,
    'mei': 5, 'may': 5,
    'jun': 6, 'juni': 6,
    'jul': 7, 'juli': 7,
    'agu': 8, 'agustus': 8, 'aug': 8,
    'sep': 9, 'september': 9, 'sept': 9,
    'okt': 10, 'oktober': 10, 'oct': 10,
    'nov': 11, 'november': 11,
    'des': 12, 'desember': 12, 'dec': 12,
}

BULAN_NAMA = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April", 5: "Mei", 6: "Juni",
    7: "Juli", 8: "Agustus", 9: "September", 10: "Oktober", 11: "November", 12: "Desember"
}

KOLOM_WAJIB = [
    "PO LN", "Komoditas", "Pemasok", "Nama Kapal", "Origin",
    "Kuantum", "Currency", "Harga", "Nilai", "Kedatangan"
]

KURS_PAJAK_URL = "https://fiskal.kemenkeu.go.id/informasi-publik/kurs-pajak"


def parse_tanggal_kedatangan(text):
    """
    Parse string rentang tanggal seperti:
      '10-15 Sep 2026'        -> 2026-09-10
      '31 Agu-4 Sep 2026'     -> 2026-08-31  (bulan pertama = Agustus)
      '5-10 Okt 2026'         -> 2026-10-05
    Aturan: yang dipakai untuk penentuan bulan & filter adalah TANGGAL PERTAMA
    beserta bulannya sendiri (bukan bulan di akhir rentang).
    Return: datetime atau None jika tidak bisa diparse / bukan format rentang tanggal.
    """
    if not text or not isinstance(text, str):
        return None
    text = text.strip()
    if not text or text.upper() in ("TBN", "-", "N/A"):
        return None

    m_year = re.search(r'(\d{4})\s*$', text)
    if not m_year:
        return None
    year = int(m_year.group(1))
    body = text[:m_year.start()].strip()
    if '-' not in body:
        return None

    left, right = body.split('-', 1)
    left, right = left.strip(), right.strip()

    m_left = re.match(r'(\d{1,2})\s*([A-Za-z]+)?', left)
    m_right = re.match(r'(\d{1,2})\s*([A-Za-z]+)?', right)
    if not m_left or not m_right:
        return None

    day1 = int(m_left.group(1))
    bulan1_text = m_left.group(2)
    bulan2_text = m_right.group(2)

    # Bulan pertama dipakai jika eksplisit tertulis; jika tidak, ikuti bulan di akhir teks
    bulan_final_text = bulan1_text if bulan1_text else bulan2_text
    if not bulan_final_text:
        return None

    bulan_key = bulan_final_text.strip().lower()
    if bulan_key not in BULAN_MAP:
        return None
    month1 = BULAN_MAP[bulan_key]

    try:
        return datetime(year, month1, day1)
    except ValueError:
        return None


def _format_id(value, decimals=0):
    """
    Format angka ke gaya Indonesia: titik sebagai pemisah ribuan, koma sebagai desimal.
    Contoh: _format_id(1234567.5, 0) -> '1.234.567'
            _format_id(40000, 2)      -> '40.000,00'
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    try:
        formatted = f"{float(value):,.{decimals}f}"
    except (ValueError, TypeError):
        return str(value)
    # Python default: ribuan="," desimal="." -> tukar jadi gaya Indonesia
    formatted = formatted.replace(",", "#").replace(".", ",").replace("#", ".")
    return formatted


def _render_preview_table_html(df, kolom_tampil):
    """
    Render tabel HTML custom untuk preview hasil konversi (read-only).
    """
    KOLOM_HIJAU = {"Bea Masuk", "PPN", "PPh", "Total"}
    KOLOM_KUANTUM = {"Kuantum"}
    KOLOM_RUPIAH = {"Nilai Invoice", "Nilai Barang", "Bea Masuk", "PPN", "PPh", "Total", "Kurs"}

    thead = "".join(
        f"<th style='padding:8px 10px; background:#f0f2f6; border:1px solid #ddd; "
        f"white-space:nowrap; font-size:12.5px;'>{col}</th>"
        for col in kolom_tampil
    )

    rows_html = []
    for _, row in df.iterrows():
        tds = []
        for col in kolom_tampil:
            val = row.get(col)
            bg = "background:#C6EFCE;" if col in KOLOM_HIJAU else ""
            align = "text-align:right;" if (col in KOLOM_RUPIAH or col in KOLOM_KUANTUM) else "text-align:left;"

            if col == "Kebutuhan Bayar" and pd.notna(val):
                dt = pd.to_datetime(val)
                text = f"{dt.month}/{dt.day}/{dt.year}"
            elif col in KOLOM_KUANTUM:
                text = _format_id(val, 2)
            elif col in KOLOM_RUPIAH:
                text = _format_id(val, 0)
            else:
                text = "" if val is None or (isinstance(val, float) and pd.isna(val)) else str(val)

            tds.append(
                f"<td style='padding:6px 10px; border:1px solid #ddd; {bg} {align} "
                f"font-size:12.5px; white-space:nowrap;'>{text}</td>"
            )
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    table_html = f"""
    <div style='overflow-x:auto; max-height:560px; overflow-y:auto; border:1px solid #ddd; border-radius:6px;'>
    <table style='border-collapse:collapse; width:100%;'>
        <thead style='position:sticky; top:0; z-index:1;'><tr>{thead}</tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
    </table>
    </div>
    """
    return table_html


def cari_sheet_data(xls_file):
    """
    Cari sheet paling relevan di workbook yang diupload: sheet dengan
    kolom wajib terbanyak yang cocok. Mengembalikan (sheet_name, header_row_idx (1-based)).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xls_file, data_only=True, read_only=True)

    kandidat = []  # list of dict: sheet, header_row, score, non_empty_rows, name_bonus

    for sn in wb.sheetnames:
        ws = wb[sn]
        best_score_sheet, best_row_sheet = -1, None
        non_empty_rows = 0
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(v is not None for v in row):
                non_empty_rows += 1
            if i <= 30:
                row_vals_str = [str(v).strip() if v is not None else "" for v in row]
                score = sum(1 for k in KOLOM_WAJIB if k in row_vals_str)
                if score > best_score_sheet:
                    best_score_sheet = score
                    best_row_sheet = i

        name_bonus = 1 if re.search(r'\b(upd|update)\b', sn, flags=re.IGNORECASE) else 0
        kandidat.append({
            "sheet": sn,
            "header_row": best_row_sheet,
            "score": best_score_sheet,
            "non_empty_rows": non_empty_rows,
            "name_bonus": name_bonus,
        })

    wb.close()

    # Filter kandidat yang minimal setengah kolom wajib cocok
    kandidat = [k for k in kandidat if k["score"] >= 5]
    if not kandidat:
        return None, None

    # Urutkan: score tertinggi > name_bonus (mengandung 'upd/update') > jumlah baris data terbanyak
    kandidat.sort(key=lambda k: (k["score"], k["name_bonus"], k["non_empty_rows"]), reverse=True)
    terpilih = kandidat[0]
    return terpilih["sheet"], terpilih["header_row"]


def load_excel_kedatangan(file_bytes, sheet_name, header_row):
    """Load sheet terpilih menjadi DataFrame bersih (drop baris kosong total)."""
    df_raw = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header_row - 1,  # pandas 0-based
        engine="openpyxl",
    )
    # Buang kolom unnamed / tanpa nama sama sekali
    df_raw = df_raw.loc[:, [c for c in df_raw.columns if not str(c).lower().startswith("unnamed")]]
    # Buang baris yang sepenuhnya kosong
    df_raw = df_raw.dropna(how="all").reset_index(drop=True)
    return df_raw


def proses_konversi(df_raw, bulan_target, tahun_target):
    """
    Terapkan aturan bisnis:
      - Skip baris dengan Origin == 'Indonesia'
      - Ambil tanggal pertama dari kolom 'Kedatangan'
    """
    hasil = []
    dilewati = []

    for idx, row in df_raw.iterrows():
        po_ln = row.get("PO LN")
        origin = row.get("Origin")
        kedatangan_raw = row.get("Kedatangan")

        komoditas = row.get("Komoditas")
        nama_kapal = row.get("Nama Kapal")
        baris_kosong_total = (
            (pd.isna(po_ln) or not str(po_ln).strip())
            and (pd.isna(komoditas) or not str(komoditas).strip())
            and (pd.isna(nama_kapal) or not str(nama_kapal).strip())
        )

        alasan_skip = None

        if baris_kosong_total:
            alasan_skip = "Baris kosong (tidak ada data)"
        elif isinstance(origin, str) and origin.strip().lower() == "indonesia":
            alasan_skip = "Origin = Indonesia"
        else:
            tgl = parse_tanggal_kedatangan(kedatangan_raw)
            if tgl is None:
                alasan_skip = "Format tanggal Kedatangan tidak valid"
            elif not (tgl.month == bulan_target and tgl.year == tahun_target):
                alasan_skip = f"Tanggal pertama di luar {BULAN_NAMA.get(bulan_target,'')} {tahun_target}"
            else:
                nilai_invoice_raw = row.get("Nilai")
                nilai_invoice = (
                    round(nilai_invoice_raw)
                    if isinstance(nilai_invoice_raw, (int, float)) and not pd.isna(nilai_invoice_raw)
                    else nilai_invoice_raw
                )
                hasil.append({
                    "PO LN": po_ln,
                    "Komoditas": row.get("Komoditas"),
                    "Pemasok": row.get("Pemasok"),
                    "Nama Kapal": row.get("Nama Kapal"),
                    "Origin": origin,
                    "Kuantum": row.get("Kuantum"),
                    "Currency": row.get("Currency"),
                    "Nilai Invoice": nilai_invoice,
                    "Kebutuhan Bayar": tgl,
                    "Kedatangan (Asli)": kedatangan_raw,
                })

        if alasan_skip:
            dilewati.append({
                "PO LN": po_ln,
                "Komoditas": row.get("Komoditas"),
                "Nama Kapal": row.get("Nama Kapal"),
                "Origin": origin,
                "Kedatangan": kedatangan_raw,
                "Alasan Dilewati": alasan_skip,
            })

    df_hasil = pd.DataFrame(hasil)
    df_dilewati = pd.DataFrame(dilewati)
    return df_hasil, df_dilewati


def hitung_kolom_pajak(df_hasil, kurs, bea_masuk_default=0, ppn_persen=11.0, pph_persen=2.5):
    """
    Tambahkan kolom perhitungan Pajak & Total
    """
    df = df_hasil.copy()
    if df.empty:
        return df

    df["Kurs"] = kurs
    df["Nilai Barang"] = df["Nilai Invoice"] * df["Kurs"]

    if "Bea Masuk" not in df.columns:
        df["Bea Masuk"] = bea_masuk_default

    df["PPN"] = df["Nilai Barang"] * (ppn_persen / 100.0)
    df["PPh"] = df["Nilai Barang"] * (pph_persen / 100.0)
    df["Total"] = df["Bea Masuk"] + df["PPN"] + df["PPh"]
    return df


def build_excel_export(df_final, bulan_target, tahun_target):
    """Susun file Excel export sesuai format contoh (judul, header, data, total baris akhir) dengan BORDER."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    kolom_export = [
        "PO LN", "Komoditas", "Pemasok", "Nama Kapal", "Origin", "Kuantum",
        "Currency", "Nilai Invoice", "Kebutuhan Bayar", "Kurs",
        "Nilai Barang", "Bea Masuk", "PPN", "PPh", "Total"
    ]

    FMT_RIBUAN_ID = '[$-421]#,##0'
    FMT_KUANTUM_ID = '[$-421]#,##0.00'

    # Konfigurasi Border Tipis
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    HIJAU_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    judul = f"RENCANA KEDATANGAN BAHAN BAKU {BULAN_NAMA.get(bulan_target,'').upper()} {tahun_target}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(kolom_export))
    ws.cell(row=1, column=1, value=judul).font = Font(bold=True, size=13, name="Arial")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True, name="Arial")

    header_row_idx = 3
    for j, col_name in enumerate(kolom_export, start=1):
        c = ws.cell(row=header_row_idx, column=j, value=col_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border  # Terapkan border ke Header

    idx_kuantum = kolom_export.index("Kuantum") + 1
    idx_nilai_barang = kolom_export.index("Nilai Barang") + 1
    idx_bea_masuk = kolom_export.index("Bea Masuk") + 1
    idx_ppn = kolom_export.index("PPN") + 1
    idx_pph = kolom_export.index("PPh") + 1
    idx_total = kolom_export.index("Total") + 1

    col_nilai_barang = get_column_letter(idx_nilai_barang)
    col_bea_masuk = get_column_letter(idx_bea_masuk)
    col_ppn = get_column_letter(idx_ppn)
    col_pph = get_column_letter(idx_pph)

    r = header_row_idx + 1
    for _, row in df_final.iterrows():
        for j, col_name in enumerate(kolom_export, start=1):
            val = row.get(col_name)

            if col_name == "Kebutuhan Bayar" and pd.notna(val):
                val = pd.to_datetime(val)
            elif col_name == "PPN":
                val = f"={col_nilai_barang}{r}*0.11"
            elif col_name == "PPh":
                val = f"={col_nilai_barang}{r}*2.5%"
            elif col_name == "Total":
                val = f"=SUM({col_bea_masuk}{r}:{col_pph}{r})"

            cell = ws.cell(row=r, column=j, value=val)
            cell.border = thin_border  # Terapkan border ke setiap sel data

            if col_name == "Kebutuhan Bayar":
                cell.number_format = "M/D/YYYY"
            elif col_name == "Kuantum":
                cell.number_format = FMT_KUANTUM_ID
            elif col_name in ("Nilai Invoice", "Nilai Barang", "Bea Masuk", "PPN", "PPh", "Total", "Kurs"):
                cell.number_format = FMT_RIBUAN_ID

            if col_name in ("Bea Masuk", "PPN", "PPh", "Total"):
                cell.fill = HIJAU_FILL

            cell.font = Font(name="Arial", size=10)
        r += 1

    # Baris Total keseluruhan dengan border penuh
    for j in range(1, len(kolom_export) + 1):
        tc = ws.cell(row=r, column=j)
        tc.border = thin_border
        if j == 1:
            tc.value = "TOTAL KESELURUHAN"
            tc.font = Font(bold=True, name="Arial")
            tc.alignment = Alignment(horizontal="center", vertical="center")
        elif j == idx_total:
            tc.value = f"=SUM({get_column_letter(idx_total)}{header_row_idx+1}:{get_column_letter(idx_total)}{r-1})"
            tc.font = Font(bold=True, name="Arial")
            tc.number_format = FMT_RIBUAN_ID
            tc.fill = HIJAU_FILL

    # Merge sel awal pada baris total untuk tampilan yang lebih rapi
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=idx_total-1)

    for j, col_name in enumerate(kolom_export, start=1):
        # Kolom angka besar (rupiah) butuh lebar ekstra supaya tidak tampil '###'
        min_width = 16 if col_name in ("Nilai Invoice", "Nilai Barang", "Bea Masuk", "PPN", "PPh", "Total", "Kurs") else 14
        ws.column_dimensions[get_column_letter(j)].width = max(min_width, len(col_name) + 4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =============================================================================
# RENDER UTAMA
# =============================================================================

def render(**kwargs):
    st.markdown("""
        <h1 style='display:flex; align-items:center; font-size:36px; margin-bottom:0;'>
            <svg xmlns="http://www.w3.org/2000/svg" width="32" height="32" fill="currentColor"
                 viewBox="0 0 16 16" style="margin-right:12px; margin-bottom:4px;">
                <path d="M8 0a1 1 0 0 1 1 1v1.076c3.416.446 6 3.36 6 6.924 0 3.866-3.134 7-7 7a6.98 6.98 0 0 1-4.988-2.09.5.5 0 0 1 .707-.708A5.98 5.98 0 0 0 8 14a6 6 0 1 0-4.242-10.243.5.5 0 0 1-.707-.707A6.97 6.97 0 0 1 7 2.076V1a1 1 0 0 1 1-1z"/>
                <path d="M7.5 5.5A.5.5 0 0 1 8 5a.5.5 0 0 1 .5.5v3.362l2.353 1.36a.5.5 0 0 1-.5.867l-2.6-1.5A.5.5 0 0 1 7.5 9.5z"/>
                <path d="M1.5 3.5A.5.5 0 0 1 2 4v1.5H3.5a.5.5 0 0 1 0 1H1.5A.5.5 0 0 1 1 6V4a.5.5 0 0 1 .5-.5"/>
            </svg>
            Rencana Kedatangan Bahan Baku
        </h1>
    """, unsafe_allow_html=True)

    st.markdown(
        "<p style='font-size:15px; opacity:0.6; margin-top:4px; margin-bottom:20px;'>"
        "Upload file Rencana Kedatangan Kapal Bahan Baku, lakukan pengecekan, lalu konversi "
        "menjadi format Rencana Kedatangan Bahan Baku per bulan yang siap diunduh."
        "</p>",
        unsafe_allow_html=True
    )

    # State persist antar-tab
    if "rkb_uploaded_bytes" not in st.session_state:
        st.session_state.rkb_uploaded_bytes = None
    if "rkb_sheet_name" not in st.session_state:
        st.session_state.rkb_sheet_name = None
    if "rkb_df_raw" not in st.session_state:
        st.session_state.rkb_df_raw = None
    if "rkb_bea_masuk_edit" not in st.session_state:
        st.session_state.rkb_bea_masuk_edit = {}

    # Menggunakan Material Icons dari Streamlit pada Tab
    tab_upload, tab_preview, tab_konversi = st.tabs([
        ":material/upload_file: Upload File Excel",
        ":material/table_view: Preview Data Excel",
        ":material/transform: Preview Hasil Konversi",
    ])

    # =========================================================================
    # TAB 1: UPLOAD
    # =========================================================================
    with tab_upload:
        st.markdown("##### Upload File Rencana Kedatangan Kapal Bahan Baku")
        st.caption(
            "Sistem akan otomatis mendeteksi sheet yang berisi data (kolom: PO LN, Komoditas, "
            "Pemasok, Nama Kapal, Origin, Kuantum, Currency, Harga, Nilai, Kedatangan, dst)."
        )

        file_upload = st.file_uploader(
            "Pilih file Excel (.xlsx)",
            type=["xlsx"],
            key="rkb_file_uploader"
        )

        if file_upload is not None:
            file_bytes = file_upload.getvalue()
            if st.button("Proses File", type="primary", icon=":material/upload_file:"):
                with st.spinner("Membaca dan mendeteksi struktur file..."):
                    try:
                        sheet_name, header_row = cari_sheet_data(io.BytesIO(file_bytes))
                        if sheet_name is None:
                            st.error(
                                "Tidak ditemukan sheet dengan struktur kolom yang sesuai "
                                "(PO LN, Komoditas, Pemasok, Nama Kapal, Origin, Kuantum, Currency, "
                                "Harga, Nilai, Kedatangan). Periksa kembali file yang diupload."
                            )
                        else:
                            df_raw = load_excel_kedatangan(file_bytes, sheet_name, header_row)
                            st.session_state.rkb_uploaded_bytes = file_bytes
                            st.session_state.rkb_sheet_name = sheet_name
                            st.session_state.rkb_df_raw = df_raw
                            st.session_state.rkb_bea_masuk_edit = {}
                            st.success(
                                f"Berhasil memuat sheet **'{sheet_name}'** — {len(df_raw)} baris data ditemukan. "
                                f"Silakan lanjut ke tab **Preview Data Excel**."
                            )
                    except Exception as e:
                        st.error(f"Gagal memproses file: {e}")

        if st.session_state.rkb_df_raw is not None:
            st.markdown("<hr style='margin:16px 0;'>", unsafe_allow_html=True)
            col_info1, col_info2 = st.columns(2)
            with col_info1:
                st.info(f":material/description: Sheet aktif: **{st.session_state.rkb_sheet_name}**")
            with col_info2:
                st.info(f":material/bar_chart: Total baris dimuat: **{len(st.session_state.rkb_df_raw)}**")

            if st.button("Hapus & Upload Ulang", icon=":material/refresh:"):
                st.session_state.rkb_uploaded_bytes = None
                st.session_state.rkb_sheet_name = None
                st.session_state.rkb_df_raw = None
                st.session_state.rkb_bea_masuk_edit = {}
                st.rerun()

    # =========================================================================
    # TAB 2: PREVIEW DATA EXCEL (SELURUH ISI)
    # =========================================================================
    with tab_preview:
        if st.session_state.rkb_df_raw is None:
            st.warning(":material/warning: Belum ada file yang diupload. Silakan upload file terlebih dahulu di tab **Upload File Excel**.")
        else:
            df_raw = st.session_state.rkb_df_raw
            st.markdown(f"##### Preview Seluruh Data — Sheet: `{st.session_state.rkb_sheet_name}`")
            st.caption(f"Menampilkan {len(df_raw)} baris data mentah hasil pembacaan file.")

            search_txt = st.text_input(
                "Cari di data (opsional)",
                placeholder="Contoh: nama kapal, pemasok, komoditas...",
                key="rkb_preview_search"
            )

            df_display = df_raw.copy()
            if search_txt.strip():
                mask = df_display.astype(str).apply(
                    lambda col: col.str.contains(search_txt, case=False, na=False)
                ).any(axis=1)
                df_display = df_display[mask]

            st.dataframe(df_display, use_container_width=True, height=520)
            st.caption(f"Menampilkan {len(df_display)} dari {len(df_raw)} baris total.")

    # =========================================================================
    # TAB 3: PREVIEW HASIL KONVERSI
    # =========================================================================
    with tab_konversi:
        if st.session_state.rkb_df_raw is None:
            st.warning(":material/warning: Belum ada file yang diupload. Silakan upload file terlebih dahulu di tab **Upload File Excel**.")
        else:
            df_raw = st.session_state.rkb_df_raw

            st.markdown("##### Filter & Parameter Konversi")

            col_bulan, col_tahun, col_kurs = st.columns([2, 1, 2])

            with col_bulan:
                bulan_pilihan = st.selectbox(
                    "Bulan Kedatangan (Kebutuhan Bayar)",
                    options=list(BULAN_NAMA.keys()),
                    format_func=lambda m: BULAN_NAMA[m],
                    index=datetime.now().month - 1,
                    key="rkb_filter_bulan"
                )
            with col_tahun:
                tahun_pilihan = st.number_input(
                    "Tahun",
                    min_value=2020, max_value=2100,
                    value=datetime.now().year,
                    step=1,
                    key="rkb_filter_tahun"
                )
            with col_kurs:
                # Menggunakan icon tautan (link) SVG sejalan dengan Bootstrap style
                st.markdown(
                    f"<label style='font-size:14px; font-weight:400;'>Kurs Pajak (KMK) "
                    f"&nbsp;<a href='{KURS_PAJAK_URL}' target='_blank' style='font-size:12px; text-decoration:none;'>"
                    f"<svg xmlns='http://www.w3.org/2000/svg' width='12' height='12' fill='currentColor' viewBox='0 0 16 16' style='vertical-align: middle; margin-right: 2px;'>"
                    f"<path d='M4.715 6.542 3.343 7.914a3 3 0 1 0 4.243 4.243l1.828-1.829A3 3 0 0 0 8.586 5.5L8 6.086a1.002 1.002 0 0 0-.154.199 2 2 0 0 1 .861 3.337L6.88 11.45a2 2 0 1 1-2.83-2.83l.793-.792a4.018 4.018 0 0 1-.128-1.287z'/>"
                    f"<path d='M6.586 4.672A3 3 0 0 0 7.414 9.5l.775-.776a2 2 0 0 1-.896-3.346L9.12 3.55a2 2 0 1 1 2.83 2.83l-.793.792c.112.42.155.855.128 1.287l1.372-1.372a3 3 0 1 0-4.243-4.243L6.586 4.672z'/>"
                    f"</svg>Cek Kurs Pajak Kemenkeu</a></label>",
                    unsafe_allow_html=True
                )
                kurs_input = st.number_input(
                    "Kurs",
                    min_value=0.0,
                    value=0.0,
                    step=1.0,
                    key="rkb_filter_kurs",
                    label_visibility="collapsed",
                    help=f"Masukkan nilai kurs pajak yang berlaku. Cek di {KURS_PAJAK_URL}"
                )

            col_ppn, col_pph = st.columns(2)
            with col_ppn:
                ppn_persen = st.number_input("PPN (%)", min_value=0.0, max_value=100.0, value=11.0, step=0.5, key="rkb_ppn")
            with col_pph:
                pph_persen = st.number_input("PPh (%)", min_value=0.0, max_value=100.0, value=2.5, step=0.5, key="rkb_pph")

            st.markdown("<hr style='margin:12px 0 16px 0;'>", unsafe_allow_html=True)

            df_hasil, df_dilewati = proses_konversi(df_raw, bulan_pilihan, int(tahun_pilihan))

            if df_hasil.empty:
                st.info(
                    f"Tidak ada data yang memenuhi kriteria untuk **{BULAN_NAMA[bulan_pilihan]} {int(tahun_pilihan)}** "
                    "(tanggal pertama Kedatangan bulan tersebut, Origin bukan Indonesia, dan PO LN terisi)."
                )
            else:
                if kurs_input <= 0:
                    st.warning(":material/warning: Masukkan nilai **Kurs** terlebih dahulu (harus lebih dari 0) untuk menghitung Nilai Barang, PPN, PPh, dan Total.")

                # Terapkan edit manual Bea Masuk yang tersimpan di session_state
                df_hasil = df_hasil.reset_index(drop=True)
                df_hasil["Bea Masuk"] = df_hasil.apply(
                    lambda r: st.session_state.rkb_bea_masuk_edit.get(f"{r['PO LN']}|{r['Nama Kapal']}", 0),
                    axis=1
                )

                df_final = hitung_kolom_pajak(
                    df_hasil, kurs=kurs_input, bea_masuk_default=0,
                    ppn_persen=ppn_persen, pph_persen=pph_persen
                )

                st.markdown(f"##### Hasil Konversi — {BULAN_NAMA[bulan_pilihan]} {int(tahun_pilihan)}")
                st.caption(
                    f"Ditemukan **{len(df_final)}** baris data. Kolom **Bea Masuk** dapat diedit manual "
                    "pada tabel input di bawah, lalu hasilnya otomatis tampil di tabel preview."
                )

                # -- Tabel kecil khusus edit Bea Masuk --
                df_bea_input = df_final[["PO LN", "Nama Kapal", "Bea Masuk"]].copy()

                edited_bea = st.data_editor(
                    df_bea_input,
                    use_container_width=True,
                    height=min(38 * (len(df_bea_input) + 1) + 3, 320),
                    disabled=["PO LN", "Nama Kapal"],
                    column_config={
                        "Bea Masuk": st.column_config.NumberColumn("Bea Masuk (bisa diedit)", format="%.0f", step=1),
                    },
                    key="rkb_data_editor_bea",
                )

                bea_masuk_changed = False
                for _, r in edited_bea.iterrows():
                    key = f"{r['PO LN']}|{r['Nama Kapal']}"
                    new_val = r["Bea Masuk"]
                    if st.session_state.rkb_bea_masuk_edit.get(key, 0) != new_val:
                        st.session_state.rkb_bea_masuk_edit[key] = new_val
                        bea_masuk_changed = True

                if bea_masuk_changed:
                    st.rerun()

                # Recompute final berdasarkan Bea Masuk terbaru untuk export
                df_export = df_final.copy()
                df_export["Bea Masuk"] = edited_bea["Bea Masuk"].values
                df_export["Nilai Barang"] = df_export["Nilai Invoice"] * df_export["Kurs"]
                df_export["PPN"] = df_export["Nilai Barang"] * (ppn_persen / 100.0)
                df_export["PPh"] = df_export["Nilai Barang"] * (pph_persen / 100.0)
                df_export["Total"] = df_export["Bea Masuk"] + df_export["PPN"] + df_export["PPh"]

                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("###### Preview Tabel Hasil")

                kolom_tampil = [
                    "PO LN", "Komoditas", "Pemasok", "Nama Kapal", "Origin", "Kuantum",
                    "Currency", "Nilai Invoice", "Kebutuhan Bayar", "Kurs",
                    "Nilai Barang", "Bea Masuk", "PPN", "PPh", "Total"
                ]
                st.markdown(
                    _render_preview_table_html(df_export, kolom_tampil),
                    unsafe_allow_html=True
                )

                total_keseluruhan = df_export["Total"].sum()
                st.markdown(
                    f"<div style='text-align:right; font-size:16px; font-weight:600; margin-top:12px;'>"
                    f"Total Keseluruhan: Rp {_format_id(total_keseluruhan, 0)}</div>",
                    unsafe_allow_html=True
                )

                st.markdown("<br>", unsafe_allow_html=True)
                excel_bytes = build_excel_export(df_export, bulan_pilihan, int(tahun_pilihan))
                nama_file = f"RENCANA_KEDATANGAN_BAHAN_BAKU_{BULAN_NAMA[bulan_pilihan]}_{int(tahun_pilihan)}.xlsx"

                st.download_button(
                    label=f"Download as Excel — {nama_file}",
                    data=excel_bytes,
                    file_name=nama_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    icon=":material/download:",
                    type="primary",
                    use_container_width=True,
                )

            if not df_dilewati.empty:
                with st.expander(f":material/search: Lihat {len(df_dilewati)} baris yang dilewati (tidak masuk hasil)"):
                    st.dataframe(df_dilewati, use_container_width=True, height=300)